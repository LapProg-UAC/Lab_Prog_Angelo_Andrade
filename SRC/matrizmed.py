import random
import json
import sys
from openpyxl import Workbook, load_workbook

from pathlib import Path

# Caminho base = pasta raiz do projeto
BASE_DIR = Path(__file__).resolve().parent.parent

# Pasta onde tens os ficheiros .txt
DATA_DIR = BASE_DIR / "data"

ficheiro_nomes = DATA_DIR / "nomesP.txt"
ficheiro_apelidos = DATA_DIR / "apelidos.txt"
ficheiro_meds = DATA_DIR / "medicamentos.txt"
EXCEL_FILE = DATA_DIR / "matrizMed.xlsx"
PRESCRICOES_FILE = DATA_DIR / "prescricoes.json"




def lertxt(filename: str) -> list[str]:
    """
    Lê um ficheiro de texto e devolve uma lista de linhas.

    Args:
        filename (str): Nome do ficheiro a ler.

    Returns:
        list[str]: Lista de linhas do ficheiro. Lista vazia em caso de erro.
    """
    try:
        with open(filename, "r", encoding="utf-8") as f:
            return [linha.strip() for linha in f]
    except FileNotFoundError:
        print(f"Erro: ficheiro '{filename}' não encontrado.")
    except Exception as e:
        print(f"Erro ao abrir '{filename}': {e}")
    return []


'''def gerarmatriz(lst:list)->list:
    matriz = []
    for i in range(len(lst) - 1):
        for j in range(len(lst) - 1):
            if i == j:
                matriz.append([[lst[i],lst[j]],0])
            else:
                matriz.append([[lst[i],lst[j]],random.randint(0,6)])
    return matriz'''


def gerar_excel(meds: list[str]) -> None:
    """
    Gera um ficheiro Excel com a matriz de interações medicamentosas.

    Args:
        meds (list[str]): Lista de medicamentos.
    """
    if not meds:
        print("Lista de medicamentos vazia. Nada a gerar.")
        return

    try:
        wb = Workbook()
        ws = wb.active

        
        ws["A1"] = "Medicamentos"
        for col, nome in enumerate(meds, start=2):
            ws.cell(row=1, column=col, value=nome)

        
        for row, nome1 in enumerate(meds, start=2):
            ws.cell(row=row, column=1, value=nome1)
            for col, nome2 in enumerate(meds, start=2):
                valor = 0 if nome1 == nome2 else random.randint(0, 6)
                ws.cell(row=row, column=col, value=valor)

        wb.save(EXCEL_FILE)
        print("Ficheiro 'matrizMed.xlsx' criado com sucesso!")

    except Exception as e:
        print(f"Erro ao criar Excel: {e}")


def valor_interacao_excel(ws, med1: str, med2: str) -> int:
    """
    Obtém o valor de interação entre dois medicamentos a partir da matriz Excel.

    Args:
        ws: Worksheet carregada.
        med1 (str): Medicamento da linha.
        med2 (str): Medicamento da coluna.

    Returns:
        int: Valor da interação.
    """
    col = None
    row = None

    
    for c in range(2, ws.max_column + 1):
        if ws.cell(row=1, column=c).value == med2:
            col = c
            break

   
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == med1:
            row = r
            break

    if col is None or row is None:
        raise ValueError(f"Medicamento não encontrado na matriz: {med1}, {med2}")

    return ws.cell(row=row, column=col).value


def calcular_balanco_excel(meds_presc: list[str], ficheiro: Path = EXCEL_FILE) -> tuple[int, list[tuple]]:
    """
    Calcula o risco total de interação entre medicamentos de uma prescrição.

    Args:
        meds_presc (list[str]): Lista de medicamentos prescritos.
        ficheiro (str): Nome do ficheiro Excel com a matriz.

    Returns:
        tuple: (risco_total, lista_de_detalhes)
    """
    wb = load_workbook(ficheiro)
    ws = wb.active

    total = 0
    detalhes = []

    for i in range(len(meds_presc)):
        for j in range(i + 1, len(meds_presc)):
            m1 = meds_presc[i]
            m2 = meds_presc[j]

            valor = valor_interacao_excel(ws, m1, m2)
            total += valor
            detalhes.append((m1, m2, valor))

    return total, detalhes


def prescricao(nomes: list[str], apelidos: list[str], meds: list[str]) -> list[dict]:
    """
    Gera prescrições médicas aleatórias para combinações de nomes e apelidos.

    Args:
        nomes (list[str]): Lista de nomes.
        apelidos (list[str]): Lista de apelidos.
        meds (list[str]): Lista de medicamentos disponíveis.

    Returns:
        list[dict]: Lista de prescrições geradas.
    """
    prescricoes = []

    for nome in nomes:
        for apelido in apelidos:
            utente_nr = random.randint(100000000, 999999999)
            nr_de_meds = random.randint(2, 4)

            meds_presc = []
            while len(meds_presc) < nr_de_meds:
                medicam = random.choice(meds)
                if medicam not in meds_presc:
                    meds_presc.append(medicam)

            total, _ = calcular_balanco_excel(meds_presc)

            nota = "Interação medicamentosa segura." if total < 15 else "Interação medicamentosa não segura."

            prescricoes.append({
                "utente": f"{nome} {apelido}",
                "numero": utente_nr,
                "medicamentos": meds_presc,
                "risco": total,
                "Nota": nota
            })

    guardar_json(prescricoes)
    return prescricoes


def guardar_json(prescricoes: list[dict], ficheiro: Path = PRESCRICOES_FILE) -> None:
    """
    Guarda as prescrições num ficheiro JSON.

    Args:
        prescricoes (list[dict]): Lista de prescrições.
        ficheiro (str): Nome do ficheiro JSON a criar.
    """
    try:
        with open(ficheiro, "w", encoding="utf-8") as f:
            json.dump(prescricoes, f, ensure_ascii=False, indent=4)
        print(f"Prescrições guardadas em '{ficheiro}'.")
    except Exception as e:
        print(f"Erro ao guardar JSON: {e}")


def main() -> None:
    """
    Função principal do programa.
    """
    nomes = lertxt(ficheiro_nomes)
    apelidos = lertxt(ficheiro_apelidos)
    meds = lertxt(ficheiro_meds)

    if not nomes or not apelidos or not meds:
        print("Erro: ficheiros de dados incompletos.")
        sys.exit(1)

    gerar_excel(meds)
    prescricao(nomes, apelidos, meds)

    print("Prescrições criadas com sucesso!")



main()
